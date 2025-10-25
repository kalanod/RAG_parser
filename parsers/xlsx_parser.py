"""Spreadsheet parser for XLSX and CSV files."""
from __future__ import annotations

from pathlib import Path
from typing import List

from core.segment import Segment
from parsers.base_parser import DocumentParser
from utils.logger import get_logger

LOGGER = get_logger(__name__)


class XLSXParser(DocumentParser):
    supported_extensions = (".xlsx", ".xls", ".csv")

    def parse(self, path: Path, **kwargs: object) -> List[Segment]:
        text = self._extract_with_pandas(path)
        if not text:
            text = self._extract_with_openpyxl(path)
        if not text:
            LOGGER.warning(
                "No spreadsheet backend available for %s. Returning empty result.",
                path,
            )
            return []
        return [Segment.from_text(text=text, source=str(path))]

    def _extract_with_pandas(self, path: Path) -> str:
        try:
            import pandas as pd  # type: ignore
        except Exception:
            return ""

        try:
            if path.suffix.lower() == ".csv":
                df = pd.read_csv(path)  # pragma: no cover
            else:
                df = pd.read_excel(path)  # pragma: no cover
        except Exception as exc:  # pragma: no cover
            LOGGER.debug("pandas failed to parse %s: %s", path, exc)
            return ""
        return df.to_csv(index=False)

    def _extract_with_openpyxl(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return ""

        wb = load_workbook(path, read_only=True, data_only=True)  # pragma: no cover
        lines = []
        for sheet in wb:
            lines.append(f"# Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_values = ["" if value is None else str(value) for value in row]
                lines.append(",".join(row_values))
        return "\n".join(lines)

